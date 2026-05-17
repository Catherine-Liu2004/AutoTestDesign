"""Sprint 2 integration tests — TestCase CRUD & Export formats."""
import json
import pytest
from unittest.mock import AsyncMock, patch
from httpx import AsyncClient, ASGITransport

from app.main import app

pytestmark = pytest.mark.asyncio


# ─── Shared fixture ───────────────────────────────────────────────────────────

@pytest.fixture
async def client():
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as c:
        yield c


MOCK_TC_RESPONSE = {
    "test_cases": [
        {
            "title": "Valid login with correct credentials",
            "preconditions": "User is registered in the system",
            "input_data": {"username": "alice", "password": "Secure@123"},
            "expected_result": "Login successful, session token returned",
            "priority": "high",
        },
        {
            "title": "Login fails with invalid password",
            "preconditions": "User exists",
            "input_data": {"username": "alice", "password": "wrong"},
            "expected_result": "HTTP 401 Unauthorized",
            "priority": "high",
        },
        {
            "title": "Login with empty username rejected",
            "preconditions": "None",
            "input_data": {"username": "", "password": "Secure@123"},
            "expected_result": "HTTP 422 Validation Error",
            "priority": "medium",
        },
    ]
}


@pytest.fixture
async def suite_with_cases(client: AsyncClient):
    """Create a requirement + test suite with mock LLM."""
    # Import
    imp = await client.post("/api/v1/requirements/import", json={
        "source_type": "direct",
        "content": "Users can log in with username and password.",
    })
    req_id = imp.json()[0]["id"]

    # Generate suite (mock LLM)
    with patch(
        "app.core.llm.LLMClient.complete_json",
        new=AsyncMock(return_value=MOCK_TC_RESPONSE),
    ):
        resp = await client.post("/api/v1/testcases/generate", json={
            "req_ids": [req_id],
            "techniques": ["equivalence_partitioning", "boundary_value_analysis", "decision_table"],
        })

    suite = resp.json()
    return suite, req_id


# ─── TestCase CRUD ────────────────────────────────────────────────────────────

async def test_list_testcases_returns_list(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    resp = await client.get("/api/v1/testcases")
    assert resp.status_code == 200
    assert isinstance(resp.json(), list)
    assert len(resp.json()) >= 3


async def test_list_testcases_filter_by_suite(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    suite_id = suite["id"]
    resp = await client.get(f"/api/v1/testcases?suite_id={suite_id}")
    assert resp.status_code == 200
    data = resp.json()
    assert all(tc["suite_id"] == suite_id for tc in data)
    assert len(data) == suite["tc_count"]


async def test_get_single_testcase(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    tc_id = suite["test_cases"][0]["id"]
    resp = await client.get(f"/api/v1/testcases/{tc_id}")
    assert resp.status_code == 200
    data = resp.json()
    assert data["id"] == tc_id
    assert data["title"] != ""
    assert data["expected_result"] != ""


async def test_get_nonexistent_testcase_returns_404(client: AsyncClient):
    resp = await client.get("/api/v1/testcases/nonexistent-id")
    assert resp.status_code == 404


async def test_update_testcase_status(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    tc_id = suite["test_cases"][0]["id"]
    resp = await client.put(f"/api/v1/testcases/{tc_id}", json={"status": "pass"})
    assert resp.status_code == 200
    assert resp.json()["status"] == "pass"


async def test_update_testcase_title_and_expected(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    tc_id = suite["test_cases"][1]["id"]
    resp = await client.put(f"/api/v1/testcases/{tc_id}", json={
        "title": "Updated: Login with correct credentials",
        "expected_result": "Returns 200 OK with Bearer token",
        "actual_result": "200 OK, token received",
    })
    assert resp.status_code == 200
    data = resp.json()
    assert data["title"] == "Updated: Login with correct credentials"
    assert data["expected_result"] == "Returns 200 OK with Bearer token"
    assert data["actual_result"] == "200 OK, token received"


async def test_delete_testcase(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    tc_id = suite["test_cases"][2]["id"]
    resp = await client.delete(f"/api/v1/testcases/{tc_id}")
    assert resp.status_code == 204
    # Confirm gone
    get_resp = await client.get(f"/api/v1/testcases/{tc_id}")
    assert get_resp.status_code == 404


async def test_all_three_techniques_in_suite(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    techniques = {tc["technique"] for tc in suite["test_cases"]}
    assert "equivalence_partitioning" in techniques
    assert "boundary_value_analysis" in techniques
    assert "decision_table" in techniques


# ─── Export tests ─────────────────────────────────────────────────────────────

async def test_export_json_format(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    suite_id = suite["id"]
    resp = await client.post("/api/v1/export", json={
        "suite_id": suite_id,
        "format": "json",
        "include": ["test_cases", "traceability_matrix"],
    })
    assert resp.status_code == 200
    assert "application/json" in resp.headers["content-type"]
    data = resp.json()
    assert data["suite_id"] == suite_id
    assert "test_cases" in data
    assert len(data["test_cases"]) >= 2  # some may have been deleted in earlier tests
    assert "traceability_matrix" in data


async def test_export_json_includes_risk_report(client: AsyncClient, suite_with_cases):
    suite, req_id = suite_with_cases
    # Analyze risk first
    with patch(
        "app.core.llm.LLMClient.complete_json",
        new=AsyncMock(return_value={"risk_score": 7.0, "priority": "high", "rationale": "Test"}),
    ):
        await client.post("/api/v1/risk/analyze", json={"req_ids": [req_id]})

    resp = await client.post("/api/v1/export", json={
        "suite_id": suite["id"],
        "format": "json",
        "include": ["test_cases", "risk_report"],
    })
    assert resp.status_code == 200
    data = resp.json()
    assert "risk_report" in data


async def test_export_excel_content_type(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    resp = await client.post("/api/v1/export", json={
        "suite_id": suite["id"],
        "format": "excel",
        "include": ["test_cases", "risk_report", "traceability_matrix"],
    })
    assert resp.status_code == 200
    assert (
        "spreadsheetml.sheet" in resp.headers["content-type"]
        or "application/octet-stream" in resp.headers["content-type"]
    )
    # Verify content is non-empty binary
    assert len(resp.content) > 1000


async def test_export_excel_valid_openpyxl(client: AsyncClient, suite_with_cases):
    """Verify Excel file can be parsed and has correct sheet/column structure."""
    import io
    import openpyxl

    suite, _ = suite_with_cases
    resp = await client.post("/api/v1/export", json={
        "suite_id": suite["id"],
        "format": "excel",
        "include": ["test_cases", "risk_report", "traceability_matrix"],
    })
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    sheet_names = wb.sheetnames
    assert "Test Cases" in sheet_names
    assert "Risk Report" in sheet_names
    assert "Traceability Matrix" in sheet_names

    # Verify Test Cases column headers
    ws_tc = wb["Test Cases"]
    headers = [cell.value for cell in next(ws_tc.iter_rows(min_row=1, max_row=1))]
    assert "TC ID" in headers
    assert "Title" in headers
    assert "Technique" in headers
    assert "Expected Result" in headers

    # Verify Risk Report headers
    ws_risk = wb["Risk Report"]
    risk_headers = [cell.value for cell in next(ws_risk.iter_rows(min_row=1, max_row=1))]
    assert "Req ID" in risk_headers
    assert "Risk Score" in risk_headers
    assert "Priority" in risk_headers

    # Verify Traceability Matrix headers
    ws_trace = wb["Traceability Matrix"]
    trace_headers = [cell.value for cell in next(ws_trace.iter_rows(min_row=1, max_row=1))]
    assert "Req ID" in trace_headers
    assert "Linked Test Cases" in trace_headers


async def test_export_nonexistent_suite_returns_404(client: AsyncClient):
    resp = await client.post("/api/v1/export", json={
        "suite_id": "nonexistent-id",
        "format": "json",
        "include": ["test_cases"],
    })
    assert resp.status_code == 404


async def test_export_unsupported_format_returns_400(client: AsyncClient, suite_with_cases):
    suite, _ = suite_with_cases
    resp = await client.post("/api/v1/export", json={
        "suite_id": suite["id"],
        "format": "pdf",
        "include": ["test_cases"],
    })
    assert resp.status_code == 422  # Pydantic validation rejects invalid enum value


# ─── TestCase filter by req_id ────────────────────────────────────────────────

async def test_list_testcases_filter_by_req_id(client: AsyncClient, suite_with_cases):
    suite, req_id = suite_with_cases
    resp = await client.get(f"/api/v1/testcases?req_id={req_id}")
    assert resp.status_code == 200
    data = resp.json()
    assert len(data) >= 1
    assert all(tc["req_id"] == req_id for tc in data)
