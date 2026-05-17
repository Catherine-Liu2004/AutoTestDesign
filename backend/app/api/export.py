import io
import json
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.models import TestCase, TestSuite, Requirement
from app.schemas.schemas import ExportRequest, TraceabilityMatrix

router = APIRouter(prefix="/export", tags=["export"])


def _build_traceability(suite: TestSuite, cases: list[TestCase]) -> dict:
    matrix: dict[str, list[str]] = {}
    for tc in cases:
        matrix.setdefault(tc.req_id, []).append(tc.id)
    return matrix


@router.post("")
async def export_artifacts(payload: ExportRequest, db: Session = Depends(get_db)):
    suite = db.query(TestSuite).filter(TestSuite.id == payload.suite_id).first()
    if not suite:
        raise HTTPException(status_code=404, detail="TestSuite not found")

    cases = db.query(TestCase).filter(TestCase.suite_id == payload.suite_id).all()
    req_ids = list({tc.req_id for tc in cases})
    reqs = db.query(Requirement).filter(Requirement.id.in_(req_ids)).all()
    req_map = {r.id: r for r in reqs}

    if payload.format == "json":
        data = {
            "suite_id": suite.id,
            "suite_name": suite.name,
        }
        if "test_cases" in payload.include:
            data["test_cases"] = [
                {
                    "id": tc.id,
                    "req_id": tc.req_id,
                    "technique": tc.technique,
                    "title": tc.title,
                    "preconditions": tc.preconditions,
                    "input_data": tc.input_data,
                    "expected_result": tc.expected_result,
                    "status": tc.status,
                    "priority": tc.priority,
                }
                for tc in cases
            ]
        if "risk_report" in payload.include:
            data["risk_report"] = [
                {
                    "req_id": r.id,
                    "risk_score": r.risk_score,
                    "priority": r.priority,
                    "rationale": r.risk_rationale,
                }
                for r in reqs
            ]
        if "traceability_matrix" in payload.include:
            data["traceability_matrix"] = _build_traceability(suite, cases)

        content = json.dumps(data, ensure_ascii=False, indent=2)
        return StreamingResponse(
            io.BytesIO(content.encode("utf-8")),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=export_{suite.id}.json"},
        )

    elif payload.format in ("excel", "csv"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        if "test_cases" in payload.include:
            ws_tc = wb.active
            ws_tc.title = "Test Cases"
            headers = [
                "TC ID", "Req ID", "Technique", "Title",
                "Preconditions", "Input Data", "Expected Result", "Status", "Priority"
            ]
            ws_tc.append(headers)
            for tc in cases:
                ws_tc.append([
                    tc.id, tc.req_id, tc.technique, tc.title,
                    tc.preconditions or "",
                    json.dumps(tc.input_data or {}, ensure_ascii=False),
                    tc.expected_result,
                    tc.status, tc.priority,
                ])

        if "risk_report" in payload.include:
            ws_risk = wb.create_sheet("Risk Report")
            ws_risk.append(["Req ID", "Raw Text", "Risk Score", "Priority", "Rationale"])
            for r in reqs:
                ws_risk.append([
                    r.id, r.raw_text[:200],
                    r.risk_score or 0.0, r.priority or "", r.risk_rationale or ""
                ])

        if "traceability_matrix" in payload.include:
            ws_trace = wb.create_sheet("Traceability Matrix")
            matrix = _build_traceability(suite, cases)
            ws_trace.append(["Req ID", "Linked Test Cases"])
            for rid, tc_ids in matrix.items():
                ws_trace.append([rid, ", ".join(tc_ids)])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=export_{suite.id}.xlsx"},
        )

    raise HTTPException(status_code=400, detail="Unsupported format")
